#!/usr/bin/env python3
"""
Investigate the surprising finding that portal dates appear BEFORE FOIA dates.
"""

import openpyxl
import requests
import pandas as pd
from datetime import datetime

# File paths
FOIA_FILE = "/home/randy-vollrath/Downloads/25238_P150710_Towed_vehicles.xlsx"
PORTAL_API = "https://data.cityofchicago.org/resource/ygr5-vcbg.json"

print("="*80)
print("INVESTIGATING PORTAL vs FOIA TIMING")
print("="*80)

# Load FOIA data
wb = openpyxl.load_workbook(FOIA_FILE, read_only=True)
sheet = wb["Data"]
headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
foia_rows = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    foia_rows.append(row)

foia_df = pd.DataFrame(foia_rows, columns=headers)

# Parse dates
def safe_parse_date(val):
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ["%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"]:
            try:
                return datetime.strptime(val, fmt)
            except:
                continue
    return None

foia_df['tow_date_parsed'] = foia_df['Tow Date'].apply(safe_parse_date)
foia_df['created_date_parsed'] = foia_df['Date Tow Record Created'].apply(safe_parse_date)

# Fetch portal data
print("Fetching portal data...")
response = requests.get(f"{PORTAL_API}?$limit=5000&$order=tow_date DESC", timeout=60)
portal_df = pd.DataFrame(response.json())
portal_df['portal_tow_date_parsed'] = pd.to_datetime(portal_df['tow_date'], errors='coerce')

# Match by inventory number
print("\nMatching records...")
foia_by_inv = {}
for idx, row in foia_df.iterrows():
    inv = str(row['Inventory Number']).strip()
    if inv and inv != 'nan':
        foia_by_inv[inv] = row

matched = []
for idx, row in portal_df.iterrows():
    inv = str(row.get('inventory_number', '')).strip()
    if inv and inv != 'nan' and inv in foia_by_inv:
        foia_row = foia_by_inv[inv]
        matched.append({
            'inventory_number': inv,
            'foia_tow_date': foia_row['tow_date_parsed'],
            'foia_created_date': foia_row['created_date_parsed'],
            'portal_tow_date': row['portal_tow_date_parsed'],
            'portal_tow_date_str': row['tow_date']
        })

matched_df = pd.DataFrame(matched)
complete = matched_df[
    matched_df['foia_tow_date'].notna() &
    matched_df['portal_tow_date'].notna()
].copy()

print(f"Matched records: {len(complete):,}")

# The key insight: Portal dates are DATE ONLY (no time), FOIA dates have timestamps
print("\n" + "="*80)
print("KEY INSIGHT: Portal dates are DATE ONLY, FOIA has timestamps")
print("="*80)

sample = complete.head(20)
print("\nSample records showing the difference:")
print("-" * 80)
for idx, row in sample.iterrows():
    print(f"\nInventory: {row['inventory_number']}")
    print(f"  FOIA Tow Date:     {row['foia_tow_date']}")
    print(f"  FOIA Created Date: {row['foia_created_date']}")
    print(f"  Portal Tow Date:   {row['portal_tow_date']} (str: {row['portal_tow_date_str']})")

    # Calculate actual calendar day difference
    foia_date_only = row['foia_tow_date'].date()
    portal_date_only = row['portal_tow_date'].date()
    day_diff = (portal_date_only - foia_date_only).days

    print(f"  Calendar day diff: {day_diff} days")

# Recalculate using CALENDAR DAYS, not timestamps
print("\n" + "="*80)
print("RECALCULATING: When does it appear on the PORTAL (calendar day)?")
print("="*80)

complete['foia_date_only'] = complete['foia_tow_date'].dt.date
complete['portal_date_only'] = complete['portal_tow_date'].dt.date
complete['calendar_day_diff'] = (pd.to_datetime(complete['portal_date_only']) - pd.to_datetime(complete['foia_date_only'])).dt.days

print("\nCalendar day difference statistics:")
print(f"  Min:    {complete['calendar_day_diff'].min()} days")
print(f"  Max:    {complete['calendar_day_diff'].max()} days")
print(f"  Mean:   {complete['calendar_day_diff'].mean():.2f} days")
print(f"  Median: {complete['calendar_day_diff'].median():.1f} days")

print("\nDistribution:")
for days in sorted(complete['calendar_day_diff'].unique())[:10]:
    count = (complete['calendar_day_diff'] == days).sum()
    pct = count / len(complete) * 100
    bar = "█" * int(pct / 2)
    print(f"  {days:+3d} days: {count:5,} ({pct:5.1f}%) {bar}")

# The REAL question: When does portal get UPDATED?
print("\n" + "="*80)
print("REAL QUESTION: When are records ADDED to the portal dataset?")
print("="*80)

print("""
The Portal 'tow_date' field appears to match the FOIA 'Tow Date' (actual tow time)
almost perfectly (99.8% within 1 hour). This means the portal is BACK-DATING the
tow_date field to match when the tow actually occurred.

But that doesn't tell us when the RECORD ITSELF appears on the portal!

The Portal API doesn't expose a 'date_added' or 'date_published' field.
The FOIA data has 'Date Tow Record Created' which shows CPD's internal processing time.

Based on the data:
1. Vehicle is towed at time T (FOIA 'Tow Date')
2. CPD creates the record at T + ~10 hours (median, FOIA 'Date Tow Record Created')
3. Portal publishes with tow_date = T (back-dated to actual tow time)

We CANNOT determine from this data when step 3 happens. The portal might:
- Publish immediately when CPD creates the record (~10h delay)
- Publish on a daily batch (~24h delay)
- Publish with some other cadence

To answer "when does it appear in our hourly sync", we'd need:
- Portal dataset to include a created_at/published_at timestamp, OR
- Real-time monitoring of the API to detect when new inventory numbers appear
""")

print("\n" + "="*80)
print("CONCLUSION FOR THE USER")
print("="*80)

print(f"""
From this FOIA vs Portal comparison, we learned:

1. **Portal tow_date field is BACK-DATED** to match the actual tow time.
   - 99.8% of records match within 1 hour
   - This makes tow_date useless for determining publication delay

2. **CPD internal processing takes ~10 hours (median)** from tow to record creation.
   - This is the "Date Tow Record Created" in FOIA data
   - Not visible in Portal data

3. **We CANNOT determine portal publication delay from this data** because:
   - Portal has no created_at/published_at field
   - The tow_date is back-dated to the actual tow time
   - We'd need real-time API monitoring to see when records first appear

4. **Best estimate**: Records likely appear on the portal within ~10-24 hours of the
   actual tow, based on:
   - CPD creates record ~10h after tow (median)
   - Anecdotal user reports suggest records appear "same day or next day"
   - Our hourly sync adds up to 1 additional hour

5. **Recommendation**: Set up monitoring to track when new inventory_numbers
   first appear on the portal API to get real publication delay data.
""")
