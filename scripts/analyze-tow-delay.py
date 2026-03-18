#!/usr/bin/env python3
"""
Cross-reference FOIA towed vehicle data with Chicago Data Portal
to determine the time gap between actual tow and portal publication.
"""

import openpyxl
import requests
import pandas as pd
from datetime import datetime, timedelta
from collections import Counter, defaultdict
import json

# File paths
FOIA_FILE = "/home/randy-vollrath/Downloads/25238_P150710_Towed_vehicles.xlsx"
PORTAL_API = "https://data.cityofchicago.org/resource/ygr5-vcbg.json"

print("="*80)
print("TOWED VEHICLE PORTAL DELAY ANALYSIS")
print("="*80)
print()

# Step 1: Load FOIA data
print("Step 1: Loading FOIA dataset...")
wb = openpyxl.load_workbook(FOIA_FILE, read_only=True)
sheet = wb["Data"]

# Read headers
headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
print(f"FOIA columns: {headers}")
print()

# Load all rows into a list
foia_rows = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    foia_rows.append(row)

print(f"Total FOIA records: {len(foia_rows):,}")

# Create DataFrame
foia_df = pd.DataFrame(foia_rows, columns=headers)

# Check for plate column and sample values
plate_columns = [col for col in headers if 'plate' in col.lower() or 'license' in col.lower()]
print(f"\nPlate-related columns: {plate_columns}")

if plate_columns:
    for col in plate_columns:
        sample_plates = foia_df[col].dropna().head(10).tolist()
        print(f"\nSample values from '{col}':")
        for plate in sample_plates[:5]:
            print(f"  {plate}")

# Identify key columns
inv_col = next((col for col in headers if 'inventory' in col.lower()), None)
tow_date_col = next((col for col in headers if col and 'tow date' in col.lower() and 'created' not in col.lower()), None)
created_col = next((col for col in headers if col and 'created' in col.lower() and 'date' in col.lower()), None)
make_col = next((col for col in headers if col and 'make' in col.lower()), None)
color_col = next((col for col in headers if col and 'color' in col.lower()), None)
pound_col = next((col for col in headers if col and 'pound' in col.lower()), None)

print(f"\nKey column mapping:")
print(f"  Inventory Number: {inv_col}")
print(f"  Tow Date: {tow_date_col}")
print(f"  Date Record Created: {created_col}")
print(f"  Make: {make_col}")
print(f"  Color: {color_col}")
print(f"  Pound: {pound_col}")

# Filter to records with inventory numbers
foia_df = foia_df[foia_df[inv_col].notna()].copy()
print(f"\nFOIA records with inventory numbers: {len(foia_df):,}")

# Parse dates
def safe_parse_date(val):
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        # Try various formats
        for fmt in ["%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"]:
            try:
                return datetime.strptime(val, fmt)
            except:
                continue
    return None

foia_df['tow_date_parsed'] = foia_df[tow_date_col].apply(safe_parse_date)
foia_df['created_date_parsed'] = foia_df[created_col].apply(safe_parse_date) if created_col else None

print(f"Successfully parsed tow dates: {foia_df['tow_date_parsed'].notna().sum():,}")
if created_col:
    print(f"Successfully parsed created dates: {foia_df['created_date_parsed'].notna().sum():,}")

# Sample FOIA records
print("\nSample FOIA records:")
sample_cols = [inv_col, tow_date_col, created_col, make_col]
sample_cols = [c for c in sample_cols if c]
print(foia_df[sample_cols].head(10).to_string())

print()
print("="*80)

# Step 2: Fetch Portal data
print("\nStep 2: Fetching Chicago Data Portal data...")
print(f"API: {PORTAL_API}")

portal_records = []
offset = 0
batch_size = 50000

while True:
    url = f"{PORTAL_API}?$limit={batch_size}&$offset={offset}&$order=tow_date DESC"
    print(f"  Fetching offset {offset}...")

    try:
        response = requests.get(url, timeout=60)
        response.raise_for_status()
        batch = response.json()

        if not batch:
            break

        portal_records.extend(batch)
        print(f"    Got {len(batch)} records (total: {len(portal_records):,})")

        if len(batch) < batch_size:
            break

        offset += batch_size

        # Safety limit
        if len(portal_records) >= 200000:
            print("    Reached 200k records, stopping.")
            break

    except Exception as e:
        print(f"    Error: {e}")
        break

print(f"\nTotal Portal records fetched: {len(portal_records):,}")

# Convert to DataFrame
portal_df = pd.DataFrame(portal_records)
print(f"Portal columns: {list(portal_df.columns)}")

# Sample portal records
print("\nSample Portal records:")
sample_portal_cols = ['inventory_number', 'tow_date', 'make', 'color', 'plate']
sample_portal_cols = [c for c in sample_portal_cols if c in portal_df.columns]
if sample_portal_cols:
    print(portal_df[sample_portal_cols].head(10).to_string())

# Parse portal tow_date
portal_df['portal_tow_date_parsed'] = pd.to_datetime(portal_df['tow_date'], errors='coerce')
print(f"\nSuccessfully parsed portal tow dates: {portal_df['portal_tow_date_parsed'].notna().sum():,}")

print()
print("="*80)

# Step 3: Match records
print("\nStep 3: Matching records by inventory_number...")

# Create lookup dictionaries
foia_by_inv = {}
for idx, row in foia_df.iterrows():
    inv = str(row[inv_col]).strip()
    if inv and inv != 'nan':
        foia_by_inv[inv] = row

portal_by_inv = {}
for idx, row in portal_df.iterrows():
    inv = str(row.get('inventory_number', '')).strip()
    if inv and inv != 'nan':
        portal_by_inv[inv] = row

print(f"FOIA unique inventory numbers: {len(foia_by_inv):,}")
print(f"Portal unique inventory numbers: {len(portal_by_inv):,}")

# Find matches
matched = []
for inv, foia_row in foia_by_inv.items():
    if inv in portal_by_inv:
        portal_row = portal_by_inv[inv]
        matched.append({
            'inventory_number': inv,
            'foia_tow_date': foia_row['tow_date_parsed'],
            'foia_created_date': foia_row.get('created_date_parsed'),
            'portal_tow_date': portal_row.get('portal_tow_date_parsed'),
            'foia_make': foia_row.get(make_col),
            'portal_make': portal_row.get('make'),
            'foia_color': foia_row.get(color_col),
            'portal_color': portal_row.get('color'),
            'pound': foia_row.get(pound_col),
        })

print(f"\nMatched records: {len(matched):,}")
print(f"Match rate: {len(matched)/len(foia_by_inv)*100:.1f}% of FOIA records")
print(f"Match rate: {len(matched)/len(portal_by_inv)*100:.1f}% of Portal records")

if not matched:
    print("\nNo matches found! Cannot continue analysis.")
    exit(1)

matched_df = pd.DataFrame(matched)

# Filter to records with all dates
complete = matched_df[
    matched_df['foia_tow_date'].notna() &
    matched_df['portal_tow_date'].notna()
].copy()

print(f"Matched records with complete dates: {len(complete):,}")

print()
print("="*80)

# Step 4: Calculate time differences
print("\nStep 4: Calculating time gaps...")

# Portal date vs FOIA tow date
complete['gap_portal_vs_tow'] = (complete['portal_tow_date'] - complete['foia_tow_date']).dt.total_seconds() / 3600
complete['gap_portal_vs_tow_hours'] = complete['gap_portal_vs_tow']

# Portal date vs FOIA created date (if available)
if 'foia_created_date' in complete.columns:
    created_available = complete[complete['foia_created_date'].notna()].copy()
    if len(created_available) > 0:
        created_available['gap_portal_vs_created'] = (
            created_available['portal_tow_date'] - created_available['foia_created_date']
        ).dt.total_seconds() / 3600
        print(f"Records with created date: {len(created_available):,}")

# FOIA created vs tow (internal CPD processing time)
if 'foia_created_date' in complete.columns:
    internal_processing = complete[complete['foia_created_date'].notna()].copy()
    if len(internal_processing) > 0:
        internal_processing['gap_created_vs_tow'] = (
            internal_processing['foia_created_date'] - internal_processing['foia_tow_date']
        ).dt.total_seconds() / 3600

# Step 5: Statistics
print("\n" + "="*80)
print("KEY FINDING: Portal tow_date vs FOIA Actual Tow Date")
print("="*80)

gaps = complete['gap_portal_vs_tow_hours'].dropna()

print(f"\nSample size: {len(gaps):,} matched records")
print(f"\nTime gap statistics (hours):")
print(f"  Min:    {gaps.min():.2f}h ({gaps.min()/24:.1f} days)")
print(f"  Max:    {gaps.max():.2f}h ({gaps.max()/24:.1f} days)")
print(f"  Mean:   {gaps.mean():.2f}h ({gaps.mean()/24:.1f} days)")
print(f"  Median: {gaps.median():.2f}h ({gaps.median()/24:.1f} days)")
print(f"  P25:    {gaps.quantile(0.25):.2f}h ({gaps.quantile(0.25)/24:.1f} days)")
print(f"  P75:    {gaps.quantile(0.75):.2f}h ({gaps.quantile(0.75)/24:.1f} days)")
print(f"  P90:    {gaps.quantile(0.90):.2f}h ({gaps.quantile(0.90)/24:.1f} days)")
print(f"  P95:    {gaps.quantile(0.95):.2f}h ({gaps.quantile(0.95)/24:.1f} days)")
print(f"  P99:    {gaps.quantile(0.99):.2f}h ({gaps.quantile(0.99)/24:.1f} days)")

# Histogram
print(f"\nDelay distribution:")
buckets = [
    ("0-1h", 0, 1),
    ("1-6h", 1, 6),
    ("6-12h", 6, 12),
    ("12-24h", 12, 24),
    ("24-48h", 24, 48),
    ("48-72h", 48, 72),
    ("3-7 days", 72, 168),
    ("7-14 days", 168, 336),
    ("14+ days", 336, float('inf'))
]

for label, min_h, max_h in buckets:
    count = ((gaps >= min_h) & (gaps < max_h)).sum()
    pct = count / len(gaps) * 100
    bar = "█" * int(pct / 2)
    print(f"  {label:12} {count:6,} ({pct:5.1f}%) {bar}")

# Check if portal date matches tow date or created date
print("\n" + "="*80)
print("DATE FIELD ALIGNMENT CHECK")
print("="*80)

# Exact matches (within 1 hour)
tow_matches = (complete['gap_portal_vs_tow_hours'].abs() < 1).sum()
print(f"\nPortal tow_date matches FOIA Tow Date (±1h): {tow_matches:,} ({tow_matches/len(complete)*100:.1f}%)")

if 'foia_created_date' in complete.columns and len(created_available) > 0:
    created_matches = (created_available['gap_portal_vs_created'].abs() < 1).sum()
    print(f"Portal tow_date matches FOIA Created Date (±1h): {created_matches:,} ({created_matches/len(created_available)*100:.1f}%)")

    print(f"\n→ The Portal tow_date appears to match the FOIA {'Tow Date' if tow_matches > created_matches else 'Created Date'}")

# Internal CPD processing time
if 'foia_created_date' in complete.columns and len(internal_processing) > 0:
    print("\n" + "="*80)
    print("INTERNAL CPD PROCESSING TIME (Created Date - Tow Date)")
    print("="*80)

    internal_gaps = internal_processing['gap_created_vs_tow'].dropna()
    if len(internal_gaps) > 0:
        print(f"\nSample size: {len(internal_gaps):,}")
        print(f"  Mean:   {internal_gaps.mean():.2f}h ({internal_gaps.mean()/24:.1f} days)")
        print(f"  Median: {internal_gaps.median():.2f}h ({internal_gaps.median()/24:.1f} days)")
        print(f"  P90:    {internal_gaps.quantile(0.90):.2f}h")

# Day of week patterns
print("\n" + "="*80)
print("DELAY PATTERNS")
print("="*80)

complete['tow_dow'] = complete['foia_tow_date'].dt.day_name()
print("\nDelay by day of week towed:")
dow_stats = complete.groupby('tow_dow')['gap_portal_vs_tow_hours'].agg(['count', 'mean', 'median'])
dow_stats = dow_stats.reindex(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'])
print(dow_stats.to_string())

# Pound patterns
if 'pound' in complete.columns:
    print("\nDelay by pound number:")
    pound_stats = complete.groupby('pound')['gap_portal_vs_tow_hours'].agg(['count', 'mean', 'median']).sort_values('count', ascending=False)
    print(pound_stats.head(10).to_string())

# Recent trend
complete_sorted = complete.sort_values('foia_tow_date')
recent = complete_sorted.tail(1000)
older = complete_sorted.head(1000)

print(f"\nRecent trend (comparing oldest 1k vs newest 1k matched records):")
print(f"  Oldest 1k mean delay: {older['gap_portal_vs_tow_hours'].mean():.1f}h ({older['gap_portal_vs_tow_hours'].mean()/24:.1f} days)")
print(f"  Newest 1k mean delay: {recent['gap_portal_vs_tow_hours'].mean():.1f}h ({recent['gap_portal_vs_tow_hours'].mean()/24:.1f} days)")

print("\n" + "="*80)
print("ANSWER TO THE REAL QUESTION")
print("="*80)

print(f"""
From the moment a car is towed, how long until it appears in our app?

1. **CPD internal processing**: ~{internal_gaps.median()/24:.1f} days (median) from tow to record creation
2. **Portal publication delay**: {gaps.median():.1f}h (median) from actual tow to portal
3. **Our hourly sync**: Up to 1 hour after portal publication

TOTAL DELAY (median case): {gaps.median():.1f}h ({gaps.median()/24:.1f} days) + up to 1h sync = ~{gaps.median()/24:.1f} days

However, the distribution shows significant variance:
- 50% of vehicles appear within {gaps.quantile(0.50)/24:.1f} days
- 90% of vehicles appear within {gaps.quantile(0.90)/24:.1f} days
- 95% of vehicles appear within {gaps.quantile(0.95)/24:.1f} days
- 5% take longer than {gaps.quantile(0.95)/24:.1f} days

The {tow_matches/len(complete)*100:.1f}% of records where portal matches FOIA tow date within 1h
suggests the portal IS trying to reflect the actual tow date, but there's
significant backend delay in the data pipeline.
""")

# Save detailed results
output_file = "/home/randy-vollrath/ticketless-chicago/tow-delay-analysis.json"
results = {
    'summary': {
        'foia_records': len(foia_df),
        'portal_records': len(portal_df),
        'matched_records': len(matched),
        'match_rate_pct': len(matched)/len(foia_by_inv)*100,
        'complete_date_records': len(complete)
    },
    'portal_vs_tow_delay_hours': {
        'min': float(gaps.min()),
        'max': float(gaps.max()),
        'mean': float(gaps.mean()),
        'median': float(gaps.median()),
        'p25': float(gaps.quantile(0.25)),
        'p75': float(gaps.quantile(0.75)),
        'p90': float(gaps.quantile(0.90)),
        'p95': float(gaps.quantile(0.95)),
        'p99': float(gaps.quantile(0.99))
    },
    'histogram': [
        {
            'bucket': label,
            'count': int(((gaps >= min_h) & (gaps < max_h)).sum()),
            'percentage': float(((gaps >= min_h) & (gaps < max_h)).sum() / len(gaps) * 100)
        }
        for label, min_h, max_h in buckets
    ],
    'date_alignment': {
        'portal_matches_foia_tow_date': int(tow_matches),
        'portal_matches_foia_tow_date_pct': float(tow_matches/len(complete)*100)
    }
}

with open(output_file, 'w') as f:
    json.dump(results, f, indent=2)

print(f"\nDetailed results saved to: {output_file}")
print("\n" + "="*80)
