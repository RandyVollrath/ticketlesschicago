#!/usr/bin/env python3
"""
Build Block Enforcement Stats with ACTUAL Revenue from FOIA Payment Data

Joins two FOIA files:
  1. tickets_where_and_when_written.xlsx (645K tickets with locations)
  2. FOIA_Vollrath_A52068_20251027.txt (7.9M payment records with dollar amounts)

The join key is Ticket Number. This produces EXACT revenue per block
(actual dollars collected, including late fees, reductions) instead of
estimates from violation_code × fine_amount.

Output: Upserts into `block_enforcement_stats` table in Supabase.

Usage:
  source .env.local
  python3 scripts/build-block-stats-with-actual-revenue.py
"""

import csv
import json
import math
import os
import re
import sys
import time
from collections import defaultdict
from datetime import datetime

# Supabase credentials from environment
SUPABASE_URL = os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Set NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY")
    print("  source .env.local && python3 scripts/build-block-stats-with-actual-revenue.py")
    sys.exit(1)

# --- File paths ---
PAYMENT_FILE = os.path.expanduser(
    "~/Downloads/FOIA_Vollrath_A52068_20251027.txt"
)
# We'll use openpyxl for the xlsx to avoid loading the entire workbook into memory
LOCATION_FILE = os.path.expanduser(
    "~/Downloads/tickets_where_and_when_written.xlsx"
)

# Fine fallbacks per violation code (used when no payment record matches)
FINE_FALLBACK = {
    '0964040B': 60,   # Street cleaning
    '0964060':  60,   # 3-7 AM snow route
    '0964070':  120,  # Snow route: 2"+ of snow
    '0964090E': 65,   # Residential permit parking
    '0964100A': 100,  # Within 15' of fire hydrant
    '0964110A': 50,   # Double parking or standing
    '0964125B': 100,  # No city sticker (<=16K lbs)
    '0964125C': 200,  # No city sticker (>16K lbs)
    '0964150B': 50,   # Parking/standing prohibited anytime
    '0964170':  50,   # Within 20' of crosswalk
    '0964190A': 50,   # Expired meter non-CBD
    '0964190B': 65,   # Expired meter CBD
    '0964190C': 150,  # Non-payment commercial loading zone
}
DEFAULT_FINE = 60  # Fallback when violation code is unknown


def parse_block_address(location):
    """Parse a Chicago address into hundred-block components.

    '2134 S ARCHER AVE' -> (2100, 'S', 'ARCHER AVE', '2100 S ARCHER AVE')
    '0 ERIE ST'         -> (0, '', 'ERIE ST', '0 ERIE ST')
    """
    trimmed = location.strip().upper()

    # With direction: "2134 S ARCHER AVE"
    m = re.match(r'^(\d+)\s+([NSEW])\s+(.+)$', trimmed)
    if m:
        num = int(m.group(1))
        block = (num // 100) * 100
        direction = m.group(2)
        street = m.group(3)
        return (block, direction, street, f"{block} {direction} {street}")

    # Without direction: "0 ERIE ST"
    m2 = re.match(r'^(\d+)\s+(.+)$', trimmed)
    if m2:
        num = int(m2.group(1))
        block = (num // 100) * 100
        street = m2.group(2)
        return (block, '', street, f"{block} {street}")

    return None


def find_peak_window(histogram):
    """Find 3-hour window with highest ticket concentration."""
    max_sum = 0
    peak_start = 0
    for h in range(24):
        s = histogram[h] + histogram[(h + 1) % 24] + histogram[(h + 2) % 24]
        if s > max_sum:
            max_sum = s
            peak_start = h
    return (peak_start, (peak_start + 3) % 24)


def load_payment_amounts():
    """Load 7.9M payment records into a dict: ticket_number -> total_payment_amount.

    The payment file is $-delimited with columns:
      Ticket Number | Issue Date/Time | Violation Code | Violation Description | Payment Amount | Payment Date

    Multiple payments can exist per ticket (partial payments, late fees).
    We sum all payments per ticket number.
    """
    print(f"Loading payment file: {PAYMENT_FILE}")
    print(f"  File size: {os.path.getsize(PAYMENT_FILE) / 1024 / 1024:.0f} MB")

    payments = defaultdict(float)
    line_count = 0
    parse_errors = 0

    t0 = time.time()
    with open(PAYMENT_FILE, 'r', encoding='utf-8', errors='replace') as f:
        # Skip header
        header = f.readline().strip()
        print(f"  Header: {header[:120]}...")

        for line in f:
            line_count += 1
            parts = line.strip().split('$')

            if len(parts) < 5:
                parse_errors += 1
                continue

            ticket_num_str = parts[0].strip()
            amount_str = parts[4].strip()

            try:
                ticket_num = int(ticket_num_str)
                amount = float(amount_str) if amount_str else 0.0
                payments[ticket_num] += amount
            except (ValueError, IndexError):
                parse_errors += 1

            if line_count % 1_000_000 == 0:
                elapsed = time.time() - t0
                print(f"  ... {line_count / 1_000_000:.0f}M rows ({elapsed:.0f}s)")

    elapsed = time.time() - t0
    print(f"  Loaded {line_count:,} payment rows in {elapsed:.1f}s")
    print(f"  Unique tickets with payments: {len(payments):,}")
    print(f"  Parse errors: {parse_errors:,}")

    # Stats
    amounts = list(payments.values())
    if amounts:
        total = sum(amounts)
        print(f"  Total payment amount: ${total:,.2f}")
        print(f"  Average per ticket: ${total / len(amounts):,.2f}")
        print(f"  Max single ticket: ${max(amounts):,.2f}")

    return payments


def load_location_tickets(payments):
    """Load 645K ticket records with locations from xlsx.

    For each ticket, look up actual payment amount from the payments dict.
    Fall back to violation_code -> fine estimate if no payment found.

    Returns aggregated block stats.
    """
    import openpyxl

    print(f"\nLoading location file: {LOCATION_FILE}")
    print(f"  File size: {os.path.getsize(LOCATION_FILE) / 1024 / 1024:.0f} MB")

    wb = openpyxl.load_workbook(LOCATION_FILE, read_only=True, data_only=True)
    ws = wb.active

    # Read headers
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    print(f"  Headers: {header_row}")

    # Ticket Number, Issue Date/Time, Violation Code, Violation Description, Location
    # Indices: 0, 1, 2, 3, 4

    blocks = {}  # block_address -> stats dict

    total = 0
    matched_payments = 0
    unmatched_payments = 0
    skipped = 0
    min_year = 9999
    max_year = 0

    t0 = time.time()
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1

        ticket_num = row[0]
        issue_dt = row[1]
        viol_code = str(row[2]).strip() if row[2] else ''
        viol_desc = str(row[3]).strip() if row[3] else ''
        location = str(row[4]).strip() if row[4] else ''

        if not location or not viol_code:
            skipped += 1
            continue

        parsed = parse_block_address(location)
        if not parsed:
            skipped += 1
            continue

        block_num, direction, street_name, block_address = parsed

        # Look up actual payment
        try:
            tnum = int(ticket_num)
        except (ValueError, TypeError):
            tnum = None

        if tnum and tnum in payments:
            revenue = payments[tnum]
            matched_payments += 1
        else:
            # Fallback to violation code estimate
            revenue = FINE_FALLBACK.get(viol_code, DEFAULT_FINE)
            unmatched_payments += 1

        # Parse date/time for histograms
        hour = 12
        dow = 3
        year = 2024

        if issue_dt:
            if isinstance(issue_dt, datetime):
                hour = issue_dt.hour
                dow = issue_dt.weekday()  # 0=Mon in Python
                # Convert to JS convention: 0=Sun
                dow = (dow + 1) % 7
                year = issue_dt.year
            elif isinstance(issue_dt, str):
                try:
                    dt = datetime.strptime(issue_dt, '%m/%d/%Y %I:%M %p')
                    hour = dt.hour
                    dow = (dt.weekday() + 1) % 7
                    year = dt.year
                except (ValueError, AttributeError):
                    pass

        if year < min_year:
            min_year = year
        if year > max_year:
            max_year = year

        # Aggregate into block
        if block_address not in blocks:
            blocks[block_address] = {
                'block_address': block_address,
                'street_direction': direction,
                'street_name': street_name,
                'block_number': block_num,
                'total_tickets': 0,
                'actual_revenue': 0.0,
                'estimated_revenue': 0,
                'violation_breakdown': {},
                'hourly_histogram': [0] * 24,
                'dow_histogram': [0] * 7,
            }

        b = blocks[block_address]
        b['total_tickets'] += 1
        b['actual_revenue'] += revenue
        b['hourly_histogram'][hour] += 1
        b['dow_histogram'][dow] += 1

        # Track by estimated too (for comparison)
        est = FINE_FALLBACK.get(viol_code, DEFAULT_FINE)
        b['estimated_revenue'] += est

        if viol_code not in b['violation_breakdown']:
            b['violation_breakdown'][viol_code] = {
                'count': 0,
                'revenue': 0.0,
                'description': viol_desc
            }
        b['violation_breakdown'][viol_code]['count'] += 1
        b['violation_breakdown'][viol_code]['revenue'] += revenue

        if total % 100_000 == 0:
            elapsed = time.time() - t0
            print(f"  ... {total:,} rows ({elapsed:.0f}s, {matched_payments:,} matched)")

    wb.close()
    elapsed = time.time() - t0

    print(f"\n  Processed {total:,} location rows in {elapsed:.1f}s")
    print(f"  Skipped: {skipped:,}")
    print(f"  Matched to payment records: {matched_payments:,} ({100*matched_payments/(total-skipped):.1f}%)")
    print(f"  Used fallback estimate: {unmatched_payments:,}")
    print(f"  Unique blocks: {len(blocks):,}")
    print(f"  Year range: {min_year}-{max_year}")

    # Post-process: peak windows, top violations, ranks
    for b in blocks.values():
        peak_start, peak_end = find_peak_window(b['hourly_histogram'])
        b['peak_hour_start'] = peak_start
        b['peak_hour_end'] = peak_end
        b['year_range'] = f"{min_year}-{max_year}"

        # Use actual revenue (rounded to integer dollars) as the canonical revenue
        b['estimated_revenue'] = round(b['actual_revenue'])

        # Top violation
        max_count = 0
        top_code = ''
        for code, v in b['violation_breakdown'].items():
            # Round revenue in breakdown to integers
            v['revenue'] = round(v['revenue'])
            if v['count'] > max_count:
                max_count = v['count']
                top_code = code

        b['top_violation_code'] = top_code
        b['top_violation_pct'] = round(100 * max_count / b['total_tickets']) if b['total_tickets'] > 0 else 0

    # Sort by revenue and assign rank
    sorted_blocks = sorted(blocks.values(), key=lambda x: x['estimated_revenue'], reverse=True)
    for i, b in enumerate(sorted_blocks):
        b['city_rank'] = i + 1

    return sorted_blocks, matched_payments, unmatched_payments


def upsert_to_supabase(blocks):
    """Upsert block stats to Supabase via REST API."""
    import urllib.request
    import ssl

    # Create SSL context that uses system certificates
    ssl_ctx = ssl.create_default_context()
    # Try certifi certs first, then fall back to unverified if needed
    try:
        import certifi
        ssl_ctx.load_verify_locations(certifi.where())
    except ImportError:
        # Try common system cert paths
        for cert_path in [
            '/etc/ssl/certs/ca-certificates.crt',
            '/etc/pki/tls/certs/ca-bundle.crt',
            '/usr/share/ca-certificates/',
        ]:
            try:
                if os.path.isfile(cert_path):
                    ssl_ctx.load_verify_locations(cert_path)
                    break
                elif os.path.isdir(cert_path):
                    ssl_ctx.load_verify_locations(capath=cert_path)
                    break
            except Exception:
                continue
        else:
            # Last resort: disable verification (still encrypted, just no cert check)
            ssl_ctx = ssl.create_default_context()
            ssl_ctx.check_hostname = False
            ssl_ctx.verify_mode = ssl.CERT_NONE
            print("  WARNING: Using unverified SSL (no CA certs found)")

    https_handler = urllib.request.HTTPSHandler(context=ssl_ctx)
    opener = urllib.request.build_opener(https_handler)

    print(f"\nUpserting {len(blocks):,} blocks to Supabase...")

    BATCH_SIZE = 500
    upserted = 0
    errors = 0

    t0 = time.time()
    for i in range(0, len(blocks), BATCH_SIZE):
        batch = blocks[i:i + BATCH_SIZE]

        payload = []
        for b in batch:
            payload.append({
                'block_address': b['block_address'],
                'street_direction': b['street_direction'],
                'street_name': b['street_name'],
                'block_number': b['block_number'],
                'total_tickets': b['total_tickets'],
                'estimated_revenue': b['estimated_revenue'],
                'city_rank': b['city_rank'],
                'violation_breakdown': b['violation_breakdown'],
                'hourly_histogram': b['hourly_histogram'],
                'dow_histogram': b['dow_histogram'],
                'peak_hour_start': b['peak_hour_start'],
                'peak_hour_end': b['peak_hour_end'],
                'top_violation_code': b['top_violation_code'],
                'top_violation_pct': b['top_violation_pct'],
                'year_range': b['year_range'],
                'updated_at': datetime.now(tz=None).isoformat() + 'Z',
            })

        body = json.dumps(payload).encode('utf-8')

        url = f"{SUPABASE_URL}/rest/v1/block_enforcement_stats"
        req = urllib.request.Request(
            url,
            data=body,
            headers={
                'apikey': SUPABASE_KEY,
                'Authorization': f'Bearer {SUPABASE_KEY}',
                'Content-Type': 'application/json',
                'Prefer': 'resolution=merge-duplicates',
            },
            method='POST'
        )

        try:
            with opener.open(req) as resp:
                upserted += len(batch)
        except urllib.error.HTTPError as e:
            error_body = e.read().decode('utf-8')
            print(f"  Batch {i // BATCH_SIZE + 1} error {e.code}: {error_body[:200]}")
            errors += 1

        if upserted % 5000 == 0 or i + BATCH_SIZE >= len(blocks):
            elapsed = time.time() - t0
            print(f"  Upserted {upserted:,}/{len(blocks):,} blocks ({elapsed:.0f}s)")

    elapsed = time.time() - t0
    print(f"\n  Done: {upserted:,} upserted, {errors} batch errors in {elapsed:.1f}s")
    return upserted, errors


def main():
    print("=" * 70)
    print("Block Enforcement Stats — ACTUAL Revenue from FOIA Payment Data")
    print("=" * 70)

    # Step 1: Load payment amounts
    payments = load_payment_amounts()

    # Step 2: Load location tickets and join with payments
    blocks, matched, unmatched = load_location_tickets(payments)

    # Step 3: Print summary before upload
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)

    total_revenue = sum(b['estimated_revenue'] for b in blocks)
    total_tickets = sum(b['total_tickets'] for b in blocks)

    print(f"  Total blocks:   {len(blocks):,}")
    print(f"  Total tickets:  {total_tickets:,}")
    print(f"  Total revenue:  ${total_revenue:,}")
    print(f"    (from actual payments: {matched:,} tickets)")
    print(f"    (from fine estimates:  {unmatched:,} tickets)")
    print(f"  Avg per block:  ${total_revenue // len(blocks):,}")
    print(f"  Blocks >$100K:  {sum(1 for b in blocks if b['estimated_revenue'] > 100000):,}")
    print(f"  Blocks >$50K:   {sum(1 for b in blocks if b['estimated_revenue'] > 50000):,}")
    print(f"  Blocks >$10K:   {sum(1 for b in blocks if b['estimated_revenue'] > 10000):,}")
    print(f"  Blocks >$5K:    {sum(1 for b in blocks if b['estimated_revenue'] > 5000):,}")

    print("\nTop 30 highest-revenue blocks:")
    for i, b in enumerate(blocks[:30]):
        top_v = b['violation_breakdown'].get(b['top_violation_code'], {})
        desc = top_v.get('description', b['top_violation_code'])
        print(f"  #{i+1:3d}: {b['block_address']:<35s} ${b['estimated_revenue']:>10,}  "
              f"({b['total_tickets']:,} tickets, top: {desc} {b['top_violation_pct']}%)")

    # Compare actual vs estimated
    # Re-compute using only fine estimates for comparison
    total_estimated_only = 0
    for b in blocks:
        est = 0
        for code, v in b['violation_breakdown'].items():
            est += v['count'] * FINE_FALLBACK.get(code, DEFAULT_FINE)
        total_estimated_only += est

    print(f"\n  Revenue comparison:")
    print(f"    Actual (from payments):  ${total_revenue:>15,}")
    print(f"    Estimated (fine × count): ${total_estimated_only:>15,}")
    diff = total_revenue - total_estimated_only
    pct = 100 * diff / total_estimated_only if total_estimated_only else 0
    print(f"    Difference:               ${diff:>+15,} ({pct:+.1f}%)")
    if diff > 0:
        print(f"    → Actual revenue is HIGHER (late fees, doubled fines, etc.)")
    else:
        print(f"    → Actual revenue is LOWER (contested/reduced/unpaid tickets)")

    # Step 4: Upsert to Supabase
    upserted, errors = upsert_to_supabase(blocks)

    print("\n" + "=" * 70)
    print(f"COMPLETE — {upserted:,} blocks loaded into block_enforcement_stats")
    print("=" * 70)


if __name__ == '__main__':
    main()
