#!/usr/bin/env python3
"""
Analyze tow record timing from FOIA data to understand:
1. How fast does CPD enter tow records (for records with real timestamps)?
2. What can we infer about midnight-stamped records?
3. What's a realistic end-to-end notification timeline?
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path
from collections import Counter, defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


def parse_excel_datetime(val):
    """Parse Excel datetime value (either datetime object or string)."""
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        # Try common formats
        for fmt in ['%Y-%m-%d %H:%M:%S', '%m/%d/%Y %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y']:
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    return None


def format_timedelta(td):
    """Format timedelta as human-readable string."""
    total_seconds = int(td.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    if hours > 0:
        return f"{hours}h {minutes}m"
    else:
        return f"{minutes}m"


def percentile(data, p):
    """Calculate percentile of sorted data."""
    if not data:
        return None
    k = (len(data) - 1) * (p / 100.0)
    f = int(k)
    c = f + 1
    if c >= len(data):
        return data[-1]
    d0 = data[f]
    d1 = data[c]
    return d0 + (d1 - d0) * (k - f)


def main():
    xlsx_path = Path("/home/randy-vollrath/Downloads/25238_P150710_Towed_vehicles.xlsx")

    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)

    print("Loading workbook...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Data"]

    # Find header row and column indices
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row[6] and "Tow Date" in str(row[6]):  # Column G
            for idx, val in enumerate(row):
                if val:
                    headers[str(val).strip()] = idx
            break

    if not headers:
        print("ERROR: Could not find header row")
        sys.exit(1)

    print(f"Found headers: {list(headers.keys())}")

    # Column indices (0-based)
    tow_date_col = headers.get("Tow Date")
    record_created_col = headers.get("Date Tow Record Created")
    reason_col = headers.get("Reason for Tow")
    type_col = headers.get("Type of Tow")
    pound_col = headers.get("Pound Number")

    if None in [tow_date_col, record_created_col]:
        print(f"ERROR: Could not find required columns")
        print(f"Tow Date: {tow_date_col}, Record Created: {record_created_col}")
        sys.exit(1)

    print("\nProcessing records...")

    # Data structures
    non_midnight_delays = []
    non_midnight_reasons = []
    midnight_records = []
    midnight_record_creation_times = []  # Just the time-of-day when records were created

    total_rows = 0
    valid_pairs = 0
    negative_delays = 0
    extreme_outliers = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1

        tow_date = parse_excel_datetime(row[tow_date_col])
        record_date = parse_excel_datetime(row[record_created_col])

        if not tow_date or not record_date:
            continue

        valid_pairs += 1

        delay = record_date - tow_date

        # Filter out negative delays
        if delay.total_seconds() < 0:
            negative_delays += 1
            continue

        # Filter out extreme outliers (>72h)
        if delay.total_seconds() > 72 * 3600:
            extreme_outliers += 1
            continue

        # Check if tow_date has a non-midnight time
        is_midnight = (tow_date.hour == 0 and tow_date.minute == 0 and tow_date.second == 0)

        if not is_midnight:
            # This is a record with a real timestamp
            non_midnight_delays.append(delay)
            if reason_col is not None:
                non_midnight_reasons.append(row[reason_col] or "Unknown")
        else:
            # Midnight-stamped record
            midnight_records.append({
                'tow_date': tow_date,
                'record_date': record_date,
                'delay': delay,
                'reason': row[reason_col] if reason_col else "Unknown",
            })
            # Extract time-of-day when the record was created
            midnight_record_creation_times.append(record_date.time())

    print(f"\nTotal rows processed: {total_rows}")
    print(f"Valid date pairs: {valid_pairs}")
    print(f"Negative delays filtered: {negative_delays}")
    print(f"Extreme outliers (>72h) filtered: {extreme_outliers}")

    # ============================================================================
    # PART 1: NON-MIDNIGHT RECORDS (Real Timestamps)
    # ============================================================================

    print("\n" + "="*80)
    print("PART 1: NON-MIDNIGHT RECORDS (Real Timestamps from CPD)")
    print("="*80)

    if non_midnight_delays:
        non_midnight_delays_sorted = sorted(non_midnight_delays, key=lambda x: x.total_seconds())
        count = len(non_midnight_delays_sorted)
        pct = (count / valid_pairs * 100) if valid_pairs > 0 else 0

        print(f"\nCount: {count:,} records ({pct:.1f}% of valid pairs)")

        print("\nCPD Record Creation Delay Distribution:")
        print(f"  Min:    {format_timedelta(non_midnight_delays_sorted[0])}")
        print(f"  P10:    {format_timedelta(timedelta(seconds=percentile([d.total_seconds() for d in non_midnight_delays_sorted], 10)))}")
        print(f"  P25:    {format_timedelta(timedelta(seconds=percentile([d.total_seconds() for d in non_midnight_delays_sorted], 25)))}")
        print(f"  Median: {format_timedelta(timedelta(seconds=percentile([d.total_seconds() for d in non_midnight_delays_sorted], 50)))}")
        print(f"  P75:    {format_timedelta(timedelta(seconds=percentile([d.total_seconds() for d in non_midnight_delays_sorted], 75)))}")
        print(f"  P90:    {format_timedelta(timedelta(seconds=percentile([d.total_seconds() for d in non_midnight_delays_sorted], 90)))}")
        print(f"  P95:    {format_timedelta(timedelta(seconds=percentile([d.total_seconds() for d in non_midnight_delays_sorted], 95)))}")
        print(f"  P99:    {format_timedelta(timedelta(seconds=percentile([d.total_seconds() for d in non_midnight_delays_sorted], 99)))}")
        print(f"  Max:    {format_timedelta(non_midnight_delays_sorted[-1])}")

        # Percentages within time windows
        print("\n% of records created within:")
        thresholds = [15*60, 30*60, 1*3600, 2*3600, 3*3600, 4*3600, 6*3600, 8*3600, 12*3600, 24*3600]
        labels = ["15 min", "30 min", "1 hour", "2 hours", "3 hours", "4 hours", "6 hours", "8 hours", "12 hours", "24 hours"]

        for threshold, label in zip(thresholds, labels):
            within = sum(1 for d in non_midnight_delays_sorted if d.total_seconds() <= threshold)
            pct = (within / count * 100)
            print(f"  {label:10s}: {pct:5.1f}%  ({within:,} records)")

        # Add 30 min sync delay for notification timeline
        print("\nWith our ~30 min average sync delay, % of users notified within:")
        SYNC_DELAY = 30 * 60  # seconds
        notification_thresholds = [1*3600, 2*3600, 3*3600, 4*3600, 6*3600]
        notification_labels = ["1 hour", "2 hours", "3 hours", "4 hours", "6 hours"]

        for threshold, label in zip(notification_thresholds, notification_labels):
            within = sum(1 for d in non_midnight_delays_sorted if d.total_seconds() + SYNC_DELAY <= threshold)
            pct = (within / count * 100)
            print(f"  {label:10s}: {pct:5.1f}%  ({within:,} users)")

        # Tow reason breakdown
        if non_midnight_reasons:
            print("\nTow Reason Breakdown (top 10):")
            reason_counts = Counter(non_midnight_reasons)
            for reason, cnt in reason_counts.most_common(10):
                pct = (cnt / count * 100)
                print(f"  {reason:40s}: {cnt:5,} ({pct:4.1f}%)")
    else:
        print("\nNo non-midnight records found.")

    # ============================================================================
    # PART 2: MIDNIGHT RECORDS - What Can We Infer?
    # ============================================================================

    print("\n" + "="*80)
    print("PART 2: MIDNIGHT RECORDS - Inference Analysis")
    print("="*80)

    if midnight_records:
        midnight_count = len(midnight_records)
        midnight_pct = (midnight_count / valid_pairs * 100) if valid_pairs > 0 else 0

        print(f"\nCount: {midnight_count:,} records ({midnight_pct:.1f}% of valid pairs)")

        # Distribution of record creation TIME OF DAY (not delay)
        print("\nRecord Creation Time-of-Day Distribution:")
        print("(When during the day does CPD enter midnight-stamped tow records?)")

        # Group by hour
        hour_counts = defaultdict(int)
        for time_obj in midnight_record_creation_times:
            hour_counts[time_obj.hour] += 1

        print("\nHourly breakdown:")
        for hour in range(24):
            cnt = hour_counts.get(hour, 0)
            pct = (cnt / midnight_count * 100) if midnight_count > 0 else 0
            bar = "█" * int(pct / 2)  # Bar chart
            print(f"  {hour:02d}:00-{hour:02d}:59  {cnt:5,} ({pct:4.1f}%)  {bar}")

        # Average/median time-of-day
        # Convert times to minutes since midnight for stats
        minutes_since_midnight = [t.hour * 60 + t.minute for t in midnight_record_creation_times]
        minutes_since_midnight.sort()

        avg_minutes = sum(minutes_since_midnight) / len(minutes_since_midnight)
        median_minutes = percentile(minutes_since_midnight, 50)

        avg_hour = int(avg_minutes // 60)
        avg_min = int(avg_minutes % 60)
        median_hour = int(median_minutes // 60)
        median_min = int(median_minutes % 60)

        print(f"\nAverage record creation time: {avg_hour:02d}:{avg_min:02d}")
        print(f"Median record creation time:  {median_hour:02d}:{median_min:02d}")

        # Speculative: if actual tow happened uniformly throughout the day
        print("\n" + "-"*80)
        print("SPECULATIVE ANALYSIS:")
        print("If we assume the ACTUAL tow happened uniformly throughout the day")
        print("(not at midnight), what would the 'true' median delay be?")
        print("-"*80)

        # For each midnight record, assume tow could have happened anytime from
        # 00:00 to 23:59 on that day. The "expected" tow time is noon (12:00).
        # Recalculate delays using noon as the tow time.

        speculative_delays = []
        for rec in midnight_records:
            # Assume tow happened at noon on the tow_date
            assumed_tow_time = rec['tow_date'].replace(hour=12, minute=0, second=0)
            speculative_delay = rec['record_date'] - assumed_tow_time
            if speculative_delay.total_seconds() >= 0:  # Only positive delays
                speculative_delays.append(speculative_delay)

        speculative_delays.sort(key=lambda x: x.total_seconds())

        if speculative_delays:
            spec_median = timedelta(seconds=percentile([d.total_seconds() for d in speculative_delays], 50))
            spec_p25 = timedelta(seconds=percentile([d.total_seconds() for d in speculative_delays], 25))
            spec_p75 = timedelta(seconds=percentile([d.total_seconds() for d in speculative_delays], 75))

            print(f"\nAssuming tow at noon on the date-stamped day:")
            print(f"  P25 delay:    {format_timedelta(spec_p25)}")
            print(f"  Median delay: {format_timedelta(spec_median)}")
            print(f"  P75 delay:    {format_timedelta(spec_p75)}")
            print("\n(This is speculative — we don't know when the actual tow occurred.)")
    else:
        print("\nNo midnight-stamped records found.")

    # ============================================================================
    # PART 3: PUTTING IT TOGETHER - Best Estimate
    # ============================================================================

    print("\n" + "="*80)
    print("PART 3: END-TO-END NOTIFICATION TIMELINE ESTIMATE")
    print("="*80)

    print("\nTimeline components:")
    print("  1. CPD enters tow record into their system")
    print("  2. Portal publishes to public API (unknown ETL delay)")
    print("  3. Our sync catches it (~30 min average, hourly checks)")
    print("  4. User gets push notification")

    if non_midnight_delays:
        # Use non-midnight data as ground truth for CPD speed
        delays_sorted = sorted(non_midnight_delays, key=lambda x: x.total_seconds())
        cpd_p50 = timedelta(seconds=percentile([d.total_seconds() for d in delays_sorted], 50))
        cpd_p25 = timedelta(seconds=percentile([d.total_seconds() for d in delays_sorted], 25))
        cpd_p75 = timedelta(seconds=percentile([d.total_seconds() for d in delays_sorted], 75))

        print("\nCPD Record Entry Speed (from non-midnight data):")
        print(f"  Fast (P25):   {format_timedelta(cpd_p25)}")
        print(f"  Typical (P50): {format_timedelta(cpd_p50)}")
        print(f"  Slow (P75):   {format_timedelta(cpd_p75)}")

        # Portal ETL delay: UNKNOWN, assume 0-30 min range
        print("\nPortal ETL Delay (CPD internal → public API):")
        print("  Unknown — assume 0-30 min range")
        print("  (The portal back-dates tow_date to match CPD's record,")
        print("   so we can't measure this delay directly)")

        # Our sync delay
        print("\nOur Sync Delay:")
        print("  Average: ~30 min (hourly checks)")
        print("  Best case: <5 min (lucky timing)")
        print("  Worst case: ~60 min (just missed a check)")

        # Calculate scenarios
        sync_avg = 30  # minutes
        sync_best = 5
        sync_worst = 60

        portal_best = 0  # minutes
        portal_likely = 15
        portal_worst = 30

        print("\n" + "-"*80)
        print("REALISTIC SCENARIOS:")
        print("-"*80)

        # Best case: fast CPD + instant portal + lucky sync
        best_total = cpd_p25.total_seconds() / 60 + portal_best + sync_best
        print(f"\nBEST CASE:")
        print(f"  CPD fast (P25): {format_timedelta(cpd_p25)}")
        print(f"  + Portal immediate: 0 min")
        print(f"  + Lucky sync: {sync_best} min")
        print(f"  = Total: ~{int(best_total)} minutes ({int(best_total/60)}h {int(best_total%60)}m)")

        # Likely case: typical CPD + some portal delay + average sync
        likely_total = cpd_p50.total_seconds() / 60 + portal_likely + sync_avg
        print(f"\nLIKELY CASE:")
        print(f"  CPD typical (P50): {format_timedelta(cpd_p50)}")
        print(f"  + Portal ETL: ~{portal_likely} min")
        print(f"  + Average sync: {sync_avg} min")
        print(f"  = Total: ~{int(likely_total)} minutes ({int(likely_total/60)}h {int(likely_total%60)}m)")

        # Worst case: slow CPD + portal batch + unlucky sync
        worst_total = cpd_p75.total_seconds() / 60 + portal_worst + sync_worst
        print(f"\nWORST CASE:")
        print(f"  CPD slow (P75): {format_timedelta(cpd_p75)}")
        print(f"  + Portal batch: ~{portal_worst} min")
        print(f"  + Unlucky sync: {sync_worst} min")
        print(f"  = Total: ~{int(worst_total)} minutes ({int(worst_total/60)}h {int(worst_total%60)}m)")

    # ============================================================================
    # PART 4: MARKETING COPY
    # ============================================================================

    print("\n" + "="*80)
    print("PART 4: HONEST MARKETING COPY")
    print("="*80)

    if non_midnight_delays:
        # Calculate key stats
        delays_sorted = sorted(non_midnight_delays, key=lambda x: x.total_seconds())
        within_2h = sum(1 for d in delays_sorted if d.total_seconds() + 30*60 <= 2*3600)
        pct_2h = (within_2h / len(delays_sorted) * 100)

        within_4h = sum(1 for d in delays_sorted if d.total_seconds() + 30*60 <= 4*3600)
        pct_4h = (within_4h / len(delays_sorted) * 100)

        median_notify = cpd_p50.total_seconds() / 60 + 30 + 15  # CPD + sync + portal

        print("\nOption 1 (Conservative):")
        print(f'  "Get notified within hours of your car being towed — typically within')
        print(f'   {int(median_notify/60)}-{int(median_notify/60)+1} hours. Most people don\'t discover their car is missing')
        print(f'   until the next day, costing ${25}/day in storage fees. Our alerts include')
        print(f'   the exact impound lot address and phone number so you can act fast."')

        print("\nOption 2 (Data-Backed):")
        print(f'  "Based on CPD records, {int(pct_2h)}% of tow notifications are sent within 2 hours')
        print(f'   and {int(pct_4h)}% within 4 hours of the tow. Even a same-day alert saves you')
        print(f'   from mounting storage fees (${25}/day) and the hassle of tracking down')
        print(f'   where your car was taken."')

        print("\nOption 3 (Benefit-Focused):")
        print(f'  "Stop discovering your towed car days later with hundreds in fees.')
        print(f'   Get same-day push notifications with the exact impound lot location,')
        print(f'   so you can retrieve your car before storage fees pile up. Most users')
        print(f'   are alerted within {int(median_notify/60)}-{int(median_notify/60)+1} hours of the tow."')

    print("\n" + "="*80)
    print("ANALYSIS COMPLETE")
    print("="*80)


if __name__ == "__main__":
    main()
