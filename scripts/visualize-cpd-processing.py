#!/usr/bin/env python3
"""
Visualize CPD internal processing time from FOIA data.
Shows how long it takes from actual tow to record creation.
"""

import openpyxl
import pandas as pd
from datetime import datetime

FOIA_FILE = "/home/randy-vollrath/Downloads/25238_P150710_Towed_vehicles.xlsx"

print("="*80)
print("CPD INTERNAL PROCESSING TIME ANALYSIS")
print("From actual tow to 'Date Tow Record Created'")
print("="*80)

# Load FOIA data
wb = openpyxl.load_workbook(FOIA_FILE, read_only=True)
sheet = wb["Data"]
headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
foia_rows = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    foia_rows.append(row)

foia_df = pd.DataFrame(foia_rows, columns=headers)

# Parse dates
def safe_parse_date(val):
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ["%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"]:
            try:
                return datetime.strptime(val, fmt)
            except:
                continue
    return None

foia_df['tow_date_parsed'] = foia_df['Tow Date'].apply(safe_parse_date)
foia_df['created_date_parsed'] = foia_df['Date Tow Record Created'].apply(safe_parse_date)

# Filter to complete records
complete = foia_df[
    foia_df['tow_date_parsed'].notna() &
    foia_df['created_date_parsed'].notna()
].copy()

print(f"\nTotal records with complete dates: {len(complete):,}")

# Calculate processing time
complete['processing_hours'] = (
    complete['created_date_parsed'] - complete['tow_date_parsed']
).dt.total_seconds() / 3600

# Statistics
print("\n" + "="*80)
print("PROCESSING TIME STATISTICS")
print("="*80)

proc_time = complete['processing_hours']

print(f"\nHow long from actual tow to CPD record creation:")
print(f"  Min:    {proc_time.min():.1f}h ({proc_time.min()/24:.1f} days)")
print(f"  Max:    {proc_time.max():.1f}h ({proc_time.max()/24:.1f} days)")
print(f"  Mean:   {proc_time.mean():.1f}h ({proc_time.mean()/24:.1f} days)")
print(f"  Median: {proc_time.median():.1f}h ({proc_time.median()/24:.1f} days)")
print(f"  P25:    {proc_time.quantile(0.25):.1f}h")
print(f"  P75:    {proc_time.quantile(0.75):.1f}h")
print(f"  P90:    {proc_time.quantile(0.90):.1f}h")
print(f"  P95:    {proc_time.quantile(0.95):.1f}h")
print(f"  P99:    {proc_time.quantile(0.99):.1f}h")

# Distribution
print(f"\n" + "="*80)
print("DISTRIBUTION")
print("="*80)

buckets = [
    ("< 1 hour", -float('inf'), 1),
    ("1-6 hours", 1, 6),
    ("6-12 hours", 6, 12),
    ("12-24 hours", 12, 24),
    ("1-2 days", 24, 48),
    ("2-3 days", 48, 72),
    ("3-7 days", 72, 168),
    ("1-2 weeks", 168, 336),
    ("2+ weeks", 336, float('inf'))
]

print(f"\nHow quickly CPD enters tows into their system:")
for label, min_h, max_h in buckets:
    count = ((proc_time >= min_h) & (proc_time < max_h)).sum()
    pct = count / len(proc_time) * 100
    bar = "█" * int(pct / 2)
    print(f"  {label:15} {count:6,} ({pct:5.1f}%) {bar}")

# Negative processing times (entered before tow date)
negative = complete[complete['processing_hours'] < 0]
if len(negative) > 0:
    print(f"\n" + "="*80)
    print("ANOMALY: Records created BEFORE tow date")
    print("="*80)
    print(f"\nCount: {len(negative):,} records ({len(negative)/len(complete)*100:.1f}%)")
    print("\nPossible explanations:")
    print("  - Scheduled/planned tows entered proactively")
    print("  - Clock skew between systems")
    print("  - Data entry errors")

    print("\nSample anomalous records:")
    sample_cols = ['Inventory Number', 'Tow Date', 'Date Tow Record Created', 'Vehicle Make']
    print(negative[sample_cols].head(10).to_string())

# Day of week patterns
print(f"\n" + "="*80)
print("PATTERNS BY DAY OF WEEK")
print("="*80)

complete['created_dow'] = complete['created_date_parsed'].dt.day_name()
print("\nProcessing time by day record was CREATED:")
dow_stats = complete.groupby('created_dow')['processing_hours'].agg(['count', 'mean', 'median'])
dow_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
dow_stats = dow_stats.reindex(dow_order)
print(dow_stats.to_string())

# Hour of day patterns
print(f"\n" + "="*80)
print("PATTERNS BY TIME OF DAY")
print("="*80)

complete['created_hour'] = complete['created_date_parsed'].dt.hour
print("\nProcessing time by hour record was CREATED:")
hour_stats = complete.groupby('created_hour')['processing_hours'].agg(['count', 'mean', 'median'])
print(hour_stats.sort_values('mean').to_string())

print(f"\n" + "="*80)
print("KEY INSIGHT")
print("="*80)

print(f"""
CPD takes a median of {proc_time.median():.1f} hours to enter a towed vehicle into
their internal system after the actual tow occurs.

{proc_time.quantile(0.50)/24:.1f} days (50th percentile)
{proc_time.quantile(0.90)/24:.1f} days (90th percentile)
{proc_time.quantile(0.95)/24:.1f} days (95th percentile)

This is the MINIMUM possible delay before a record could appear on the
Chicago Data Portal. The actual portal publication delay is unknown
(portal doesn't expose a created_at timestamp).

Our best estimate: vehicles appear in our app within 10-24 hours of
being towed, based on this CPD processing time plus assumed portal ETL delay.
""")
