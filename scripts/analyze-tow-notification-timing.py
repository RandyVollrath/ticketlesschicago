#!/usr/bin/env python3
"""
Analyze FOIA towed vehicles data to determine notification timing windows.
Calculates what percentage of tows would have records available for notification
within various time windows after the actual tow event.
"""

import sys
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook

def parse_datetime(cell_value):
    """Parse datetime from Excel cell value."""
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value
    if isinstance(cell_value, str):
        # Try common formats
        for fmt in ['%m/%d/%Y %H:%M', '%Y-%m-%d %H:%M:%S', '%m/%d/%Y %I:%M:%S %p']:
            try:
                return datetime.strptime(cell_value.strip(), fmt)
            except ValueError:
                continue
    return None

def format_timedelta(td):
    """Format timedelta as human-readable string."""
    total_seconds = int(td.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    if hours > 0:
        return f"{hours}h {minutes}m"
    else:
        return f"{minutes}m"

def get_hour_bucket(dt):
    """Get hour-of-day bucket (0-23)."""
    return dt.hour

def main():
    xlsx_path = '/home/randy-vollrath/Downloads/25238_P150710_Towed_vehicles.xlsx'

    print("Loading Excel file...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = wb['Data']

    # Find header row and column indices
    headers = None
    header_row_idx = None
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if 'Tow Date' in row:
            headers = row
            header_row_idx = idx
            break

    if not headers:
        print("ERROR: Could not find header row with 'Tow Date'")
        sys.exit(1)

    tow_date_col = None
    record_created_col = None
    tow_reason_col = None

    for idx, header in enumerate(headers):
        if header == 'Tow Date':
            tow_date_col = idx
        elif header == 'Date Tow Record Created':
            record_created_col = idx
        elif header == 'Reason for Tow':
            tow_reason_col = idx

    if tow_date_col is None or record_created_col is None:
        print(f"ERROR: Missing required columns. Found: {headers}")
        sys.exit(1)

    print(f"Found headers at row {header_row_idx}")
    print(f"  Tow Date: column {tow_date_col}")
    print(f"  Date Tow Record Created: column {record_created_col}")
    print(f"  Tow Reason: column {tow_reason_col if tow_reason_col is not None else 'NOT FOUND'}")

    # Parse data
    delays = []
    delays_with_metadata = []  # (delay_seconds, tow_datetime, tow_reason)
    skipped_no_dates = 0
    skipped_negative = 0
    skipped_outlier = 0

    print("\nParsing data rows...")
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if len(row) <= max(tow_date_col, record_created_col):
            continue

        tow_date = parse_datetime(row[tow_date_col])
        record_created = parse_datetime(row[record_created_col])
        tow_reason = row[tow_reason_col] if tow_reason_col is not None and len(row) > tow_reason_col else None

        if not tow_date or not record_created:
            skipped_no_dates += 1
            continue

        delay = record_created - tow_date
        delay_seconds = delay.total_seconds()

        # Filter negative delays
        if delay_seconds < 0:
            skipped_negative += 1
            continue

        # Filter extreme outliers (>72 hours)
        if delay_seconds > 72 * 3600:
            skipped_outlier += 1
            continue

        delays.append(delay_seconds)
        delays_with_metadata.append((delay_seconds, tow_date, tow_reason))

    wb.close()

    if not delays:
        print("ERROR: No valid data found")
        sys.exit(1)

    total_records = len(delays)
    print(f"\nData loaded:")
    print(f"  Valid records: {total_records:,}")
    print(f"  Skipped (no dates): {skipped_no_dates:,}")
    print(f"  Skipped (negative delay): {skipped_negative:,}")
    print(f"  Skipped (>72h outlier): {skipped_outlier:,}")

    # Sort delays for percentile calculations
    delays_sorted = sorted(delays)

    # Calculate percentiles
    def percentile(data, pct):
        idx = int(len(data) * pct)
        return data[min(idx, len(data) - 1)]

    median = percentile(delays_sorted, 0.5)
    p75 = percentile(delays_sorted, 0.75)
    p90 = percentile(delays_sorted, 0.90)
    p95 = percentile(delays_sorted, 0.95)

    print(f"\nOverall statistics:")
    print(f"  Median delay: {format_timedelta(timedelta(seconds=median))}")
    print(f"  75th percentile: {format_timedelta(timedelta(seconds=p75))}")
    print(f"  90th percentile: {format_timedelta(timedelta(seconds=p90))}")
    print(f"  95th percentile: {format_timedelta(timedelta(seconds=p95))}")

    # Fine-grained early window analysis
    print("\n" + "="*80)
    print("RECORD CREATION TIMING (from actual tow to record in CPD system)")
    print("="*80)

    windows = [
        (30 * 60, "30 minutes"),
        (1 * 3600, "1 hour"),
        (1.5 * 3600, "1.5 hours"),
        (2 * 3600, "2 hours"),
        (3 * 3600, "3 hours"),
        (4 * 3600, "4 hours"),
        (5 * 3600, "5 hours"),
        (6 * 3600, "6 hours"),
        (8 * 3600, "8 hours"),
        (12 * 3600, "12 hours"),
        (24 * 3600, "24 hours"),
    ]

    for window_seconds, window_label in windows:
        count = sum(1 for d in delays if d <= window_seconds)
        pct = (count / total_records) * 100
        print(f"  Within {window_label:12s}: {count:6,} / {total_records:,} ({pct:5.1f}%)")

    # Notification delay analysis (record creation + sync delay)
    print("\n" + "="*80)
    print("USER NOTIFICATION TIMING (record creation + hourly sync delay)")
    print("Assumes portal publishes immediately, sync adds avg 30min (0-60min uniform)")
    print("="*80)

    notification_windows = [
        (2 * 3600, "2 hours", 1 * 3600),      # 2h notification needs ~1h record creation
        (3 * 3600, "3 hours", 2 * 3600),      # 3h notification needs ~2h record creation
        (4 * 3600, "4 hours", 3 * 3600),      # 4h notification needs ~3h record creation
        (6 * 3600, "6 hours", 5 * 3600),      # 6h notification needs ~5h record creation
        (8 * 3600, "8 hours", 7 * 3600),      # 8h notification needs ~7h record creation
        (12 * 3600, "12 hours", 11 * 3600),   # 12h notification needs ~11h record creation
        (24 * 3600, "24 hours", 23 * 3600),   # 24h notification needs ~23h record creation
    ]

    for notif_window, notif_label, record_threshold in notification_windows:
        # Count records created within threshold (leaving room for sync delay)
        count = sum(1 for d in delays if d <= record_threshold)
        pct = (count / total_records) * 100
        print(f"  Notified within {notif_label:8s}: {count:6,} / {total_records:,} ({pct:5.1f}%)")

    # Time of day analysis
    print("\n" + "="*80)
    print("RECORD CREATION SPEED BY TIME OF DAY (hour tow occurred)")
    print("="*80)

    hourly_delays = defaultdict(list)
    for delay_sec, tow_dt, _ in delays_with_metadata:
        hour = get_hour_bucket(tow_dt)
        hourly_delays[hour].append(delay_sec)

    print(f"\n{'Hour':6s}  {'Count':>6s}  {'Median':>10s}  {'<2h':>6s}  {'<4h':>6s}  {'<6h':>6s}")
    print("-" * 60)

    for hour in range(24):
        if hour not in hourly_delays:
            continue

        hour_data = sorted(hourly_delays[hour])
        count = len(hour_data)
        median_sec = percentile(hour_data, 0.5)
        within_2h = sum(1 for d in hour_data if d <= 2 * 3600)
        within_4h = sum(1 for d in hour_data if d <= 4 * 3600)
        within_6h = sum(1 for d in hour_data if d <= 6 * 3600)

        pct_2h = (within_2h / count) * 100
        pct_4h = (within_4h / count) * 100
        pct_6h = (within_6h / count) * 100

        hour_label = f"{hour:02d}:00"
        median_label = format_timedelta(timedelta(seconds=median_sec))

        print(f"{hour_label:6s}  {count:6,}  {median_label:>10s}  {pct_2h:5.1f}%  {pct_4h:5.1f}%  {pct_6h:5.1f}%")

    # Tow reason analysis
    if tow_reason_col is not None:
        print("\n" + "="*80)
        print("RECORD CREATION SPEED BY TOW REASON (top 10 by volume)")
        print("="*80)

        reason_delays = defaultdict(list)
        for delay_sec, _, tow_reason in delays_with_metadata:
            if tow_reason:
                reason_delays[str(tow_reason).strip()].append(delay_sec)

        # Sort by volume
        reasons_by_volume = sorted(reason_delays.items(), key=lambda x: len(x[1]), reverse=True)

        print(f"\n{'Tow Reason':40s}  {'Count':>6s}  {'Median':>10s}  {'<4h':>6s}")
        print("-" * 80)

        for reason, reason_data in reasons_by_volume[:10]:
            count = len(reason_data)
            reason_sorted = sorted(reason_data)
            median_sec = percentile(reason_sorted, 0.5)
            within_4h = sum(1 for d in reason_data if d <= 4 * 3600)
            pct_4h = (within_4h / count) * 100

            reason_truncated = reason[:38] if len(reason) > 38 else reason
            median_label = format_timedelta(timedelta(seconds=median_sec))

            print(f"{reason_truncated:40s}  {count:6,}  {median_label:>10s}  {pct_4h:5.1f}%")

    print("\n" + "="*80)
    print("ANALYSIS COMPLETE")
    print("="*80)

if __name__ == '__main__':
    main()
