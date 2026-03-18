#!/usr/bin/env python3
"""
Tow Data Analysis Script
Part 1: Verify "93% within 15 minutes" claim from CPD FOIA data
Part 2: Analyze OEMC dispatch data and cross-reference
"""

import csv
from datetime import datetime, timedelta
from openpyxl import load_workbook
import random
from collections import defaultdict

print("=" * 80)
print("TOW DATA ANALYSIS")
print("=" * 80)

# ============================================================================
# PART 1: CPD FOIA DATA - Verify "93% within 15 minutes" claim
# ============================================================================

print("\n" + "=" * 80)
print("PART 1: CPD FOIA DATA ANALYSIS")
print("=" * 80)

xlsx_path = "/home/randy-vollrath/Downloads/25238_P150710_Towed_vehicles.xlsx"
print(f"\nLoading: {xlsx_path}")

wb = load_workbook(xlsx_path, data_only=True)
sheet = wb["Data"]

# Get headers
headers = []
for cell in sheet[1]:
    headers.append(cell.value)

print(f"\nColumns found: {len(headers)}")
print("Column G (index 6):", headers[6] if len(headers) > 6 else "NOT FOUND")
print("Column Q (index 16):", headers[16] if len(headers) > 16 else "NOT FOUND")

# Find column indices
tow_date_idx = 6  # Column G
record_created_idx = 16  # Column Q
reason_idx = headers.index("Reason for Tow") if "Reason for Tow" in headers else None

print(f"\nReason for Tow column index: {reason_idx}")

# Read all data
all_rows = []
midnight_rows = []
non_midnight_rows = []

print("\nProcessing rows...")
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    if row_idx % 10000 == 0:
        print(f"  Processed {row_idx} rows...")

    tow_date = row[tow_date_idx]
    record_created = row[record_created_idx]

    if not tow_date or not record_created:
        continue

    # Convert to datetime if needed
    if isinstance(tow_date, str):
        try:
            tow_date = datetime.strptime(tow_date, "%Y-%m-%d %H:%M:%S")
        except:
            try:
                tow_date = datetime.strptime(tow_date, "%m/%d/%Y %H:%M:%S")
            except:
                continue

    if isinstance(record_created, str):
        try:
            record_created = datetime.strptime(record_created, "%Y-%m-%d %H:%M:%S")
        except:
            try:
                record_created = datetime.strptime(record_created, "%m/%d/%Y %H:%M:%S")
            except:
                continue

    # Check if midnight (00:00:00)
    is_midnight = tow_date.hour == 0 and tow_date.minute == 0 and tow_date.second == 0

    row_data = {
        'tow_date': tow_date,
        'record_created': record_created,
        'reason': row[reason_idx] if reason_idx else None,
        'row': row
    }

    all_rows.append(row_data)

    if is_midnight:
        midnight_rows.append(row_data)
    else:
        non_midnight_rows.append(row_data)

print(f"\nTotal rows with valid data: {len(all_rows)}")
print(f"Midnight timestamps (00:00:00): {len(midnight_rows)} ({len(midnight_rows)*100/len(all_rows):.1f}%)")
print(f"Non-midnight timestamps: {len(non_midnight_rows)} ({len(non_midnight_rows)*100/len(all_rows):.1f}%)")

# Analyze non-midnight records
print("\n" + "-" * 80)
print("NON-MIDNIGHT TIMESTAMP ANALYSIS")
print("-" * 80)

# Calculate delays
delay_buckets = {
    '0-1min': [],
    '1-5min': [],
    '5-15min': [],
    '15-30min': [],
    '30min-1h': [],
    '1-2h': [],
    '2-4h': [],
    '4-8h': [],
    '8-24h': [],
    '24h+': []
}

exact_matches = 0

for row_data in non_midnight_rows:
    delay = row_data['record_created'] - row_data['tow_date']
    delay_seconds = delay.total_seconds()
    delay_minutes = delay_seconds / 60

    row_data['delay'] = delay
    row_data['delay_minutes'] = delay_minutes

    # Check for exact match
    if delay_seconds == 0:
        exact_matches += 1

    # Bucket it
    if delay_minutes < 1:
        delay_buckets['0-1min'].append(row_data)
    elif delay_minutes < 5:
        delay_buckets['1-5min'].append(row_data)
    elif delay_minutes < 15:
        delay_buckets['5-15min'].append(row_data)
    elif delay_minutes < 30:
        delay_buckets['15-30min'].append(row_data)
    elif delay_minutes < 60:
        delay_buckets['30min-1h'].append(row_data)
    elif delay_minutes < 120:
        delay_buckets['1-2h'].append(row_data)
    elif delay_minutes < 240:
        delay_buckets['2-4h'].append(row_data)
    elif delay_minutes < 480:
        delay_buckets['4-8h'].append(row_data)
    elif delay_minutes < 1440:
        delay_buckets['8-24h'].append(row_data)
    else:
        delay_buckets['24h+'].append(row_data)

print(f"\nDELAY DISTRIBUTION (n={len(non_midnight_rows)}):")
print("-" * 80)

cumulative = 0
for bucket_name in ['0-1min', '1-5min', '5-15min', '15-30min', '30min-1h', '1-2h', '2-4h', '4-8h', '8-24h', '24h+']:
    count = len(delay_buckets[bucket_name])
    pct = count * 100 / len(non_midnight_rows)
    cumulative += count
    cumulative_pct = cumulative * 100 / len(non_midnight_rows)
    print(f"{bucket_name:12} {count:6} ({pct:5.2f}%)  |  Cumulative: {cumulative:6} ({cumulative_pct:5.2f}%)")

print(f"\nExact matches (0 second delay): {exact_matches} ({exact_matches*100/len(non_midnight_rows):.2f}%)")

# Check the <15min claim
under_15min = len(delay_buckets['0-1min']) + len(delay_buckets['1-5min']) + len(delay_buckets['5-15min'])
print(f"\n✓ Records within 15 minutes: {under_15min} ({under_15min*100/len(non_midnight_rows):.2f}%)")

# Sample rows from each bucket
print("\n" + "-" * 80)
print("RANDOM SAMPLE ROWS (20 total)")
print("-" * 80)

samples = []
samples.extend(random.sample(delay_buckets['0-1min'], min(5, len(delay_buckets['0-1min']))))
samples.extend(random.sample(delay_buckets['5-15min'], min(5, len(delay_buckets['5-15min']))))
if delay_buckets['30min-1h'] or delay_buckets['1-2h'] or delay_buckets['2-4h']:
    combined_1_4h = delay_buckets['30min-1h'] + delay_buckets['1-2h'] + delay_buckets['2-4h']
    samples.extend(random.sample(combined_1_4h, min(5, len(combined_1_4h))))
if delay_buckets['4-8h'] or delay_buckets['8-24h'] or delay_buckets['24h+']:
    combined_4h_plus = delay_buckets['4-8h'] + delay_buckets['8-24h'] + delay_buckets['24h+']
    samples.extend(random.sample(combined_4h_plus, min(5, len(combined_4h_plus))))

print(f"\n{'Tow Date':<20} | {'Record Created':<20} | {'Delay':<12} | Reason")
print("-" * 120)
for sample in samples:
    delay_str = f"{sample['delay_minutes']:.1f}min"
    if sample['delay_minutes'] > 60:
        delay_str = f"{sample['delay_minutes']/60:.1f}h"
    print(f"{sample['tow_date']!s:<20} | {sample['record_created']!s:<20} | {delay_str:<12} | {sample['reason']}")

# Examine 0-1 minute records in detail
print("\n" + "-" * 80)
print("0-1 MINUTE DELAY RECORDS - EXAMINING SECONDS")
print("-" * 80)

zero_to_one_min = delay_buckets['0-1min']
print(f"\nTotal 0-1 minute records: {len(zero_to_one_min)}")

if len(zero_to_one_min) > 0:
    print("\n10 sample records showing full datetime with seconds:")
    print(f"{'Tow Date (full)':<25} | {'Record Created (full)':<25} | {'Diff (sec)':<12}")
    print("-" * 80)

    for sample in random.sample(zero_to_one_min, min(10, len(zero_to_one_min))):
        diff_sec = (sample['record_created'] - sample['tow_date']).total_seconds()
        print(f"{sample['tow_date']!s:<25} | {sample['record_created']!s:<25} | {diff_sec:<12.0f}")

# Compare midnight vs non-midnight groups
print("\n" + "-" * 80)
print("MIDNIGHT vs NON-MIDNIGHT COMPARISON")
print("-" * 80)

print("\nTow reasons in non-midnight group:")
non_midnight_reasons = defaultdict(int)
for row in non_midnight_rows[:1000]:  # Sample first 1000
    if row['reason']:
        non_midnight_reasons[row['reason']] += 1

for reason, count in sorted(non_midnight_reasons.items(), key=lambda x: x[1], reverse=True)[:10]:
    print(f"  {reason}: {count}")

print("\nTow reasons in midnight group:")
midnight_reasons = defaultdict(int)
for row in midnight_rows[:1000]:  # Sample first 1000
    if row['reason']:
        midnight_reasons[row['reason']] += 1

for reason, count in sorted(midnight_reasons.items(), key=lambda x: x[1], reverse=True)[:10]:
    print(f"  {reason}: {count}")

# ============================================================================
# PART 2: OEMC DISPATCH DATA
# ============================================================================

print("\n\n" + "=" * 80)
print("PART 2: OEMC DISPATCH DATA ANALYSIS")
print("=" * 80)

csv_path = "/home/randy-vollrath/Downloads/FOIA_F261085_TOW_AUG_OCT_2025.csv"
print(f"\nLoading: {csv_path}")

# First check what delimiter is used
with open(csv_path, 'r', encoding='utf-8-sig') as f:
    first_line = f.readline()
    print(f"\nFirst line raw: {repr(first_line[:200])}")

# The file appears to use | as delimiter
with open(csv_path, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f, delimiter='|')

    # Get column names
    columns = reader.fieldnames
    print(f"\nColumns found ({len(columns)}):")
    for i, col in enumerate(columns, 1):
        print(f"  {i}. {col}")

    # Read all rows
    oemc_rows = [dict(row) for row in reader]

print(f"\nTotal rows: {len(oemc_rows)}")

# Print 10 sample rows
print("\n10 Sample Rows:")
print("-" * 80)
for i, row in enumerate(oemc_rows[:10], 1):
    print(f"\nRow {i}:")
    for key, value in row.items():
        if value and value.strip():
            print(f"  {key}: {value}")

# Analyze datetime columns
print("\n" + "-" * 80)
print("DATETIME COLUMN ANALYSIS")
print("-" * 80)

datetime_cols = []
for col in columns:
    col_lower = col.lower()
    if any(word in col_lower for word in ['date', 'time', 'created', 'dispatch', 'close', 'complete']):
        datetime_cols.append(col)

print(f"\nPotential datetime columns: {datetime_cols}")

for col in datetime_cols:
    values = [row[col] for row in oemc_rows if row[col] and row[col].strip()]
    if values:
        print(f"\n{col}:")
        print(f"  Sample values: {values[:3]}")
        print(f"  Non-empty count: {len(values)}")

        # Try to parse and get min/max
        parsed_dates = []
        for val in values:
            try:
                # Try common formats
                for fmt in ["%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"]:
                    try:
                        dt = datetime.strptime(val, fmt)
                        parsed_dates.append(dt)
                        break
                    except:
                        continue
            except:
                pass

        if parsed_dates:
            print(f"  Min: {min(parsed_dates)}")
            print(f"  Max: {max(parsed_dates)}")
            print(f"  Successfully parsed: {len(parsed_dates)}/{len(values)}")

# Check for cross-referenceable fields
print("\n" + "-" * 80)
print("CROSS-REFERENCE POTENTIAL")
print("-" * 80)

print("\nLooking for fields that could match CPD FOIA data...")
potential_match_fields = []

for col in columns:
    col_lower = col.lower()
    if any(word in col_lower for word in ['inventory', 'event', 'number', 'location', 'address', 'plate', 'license']):
        potential_match_fields.append(col)
        sample_vals = [row[col] for row in oemc_rows[:10] if row[col] and row[col].strip()]
        print(f"\n{col}:")
        print(f"  Sample values: {sample_vals[:5]}")

# Calculate dispatch to completion duration if applicable
print("\n" + "-" * 80)
print("DISPATCH TO COMPLETION DURATION")
print("-" * 80)

# Look for dispatch and completion time columns
dispatch_col = None
completion_col = None

for col in columns:
    col_lower = col.lower()
    if 'dispatch' in col_lower:
        dispatch_col = col
    if 'close' in col_lower:
        completion_col = col

print(f"\nDispatch column: {dispatch_col}")
print(f"Completion column: {completion_col}")

if dispatch_col and completion_col:
    durations = []

    for row in oemc_rows:
        dispatch_val = row.get(dispatch_col, '').strip()
        completion_val = row.get(completion_col, '').strip()

        if not dispatch_val or not completion_val:
            continue

        try:
            for fmt in ["%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"]:
                try:
                    dispatch_dt = datetime.strptime(dispatch_val, fmt)
                    completion_dt = datetime.strptime(completion_val, fmt)
                    duration = completion_dt - dispatch_dt
                    durations.append(duration.total_seconds() / 60)  # minutes
                    break
                except:
                    continue
        except:
            pass

    if durations:
        print(f"\nDuration statistics (n={len(durations)}):")
        durations.sort()
        print(f"  Min: {min(durations):.1f} minutes")
        print(f"  Max: {max(durations):.1f} minutes")
        print(f"  Median: {durations[len(durations)//2]:.1f} minutes")
        print(f"  Mean: {sum(durations)/len(durations):.1f} minutes")

        # Distribution
        duration_buckets = {
            '0-15min': 0,
            '15-30min': 0,
            '30-60min': 0,
            '1-2h': 0,
            '2-4h': 0,
            '4h+': 0
        }

        for d in durations:
            if d < 15:
                duration_buckets['0-15min'] += 1
            elif d < 30:
                duration_buckets['15-30min'] += 1
            elif d < 60:
                duration_buckets['30-60min'] += 1
            elif d < 120:
                duration_buckets['1-2h'] += 1
            elif d < 240:
                duration_buckets['2-4h'] += 1
            else:
                duration_buckets['4h+'] += 1

        print("\n  Distribution:")
        for bucket, count in duration_buckets.items():
            pct = count * 100 / len(durations)
            print(f"    {bucket:12} {count:6} ({pct:5.2f}%)")
else:
    print("\n⚠ Could not find both dispatch and completion columns")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
